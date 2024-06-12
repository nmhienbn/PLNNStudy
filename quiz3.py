from docx import Document
from docx.oxml.ns import qn
import pandas as pd
import random
import os
from openpyxl import load_workbook
from tkinter import *
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk


class QuizApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Quiz Application")
        self.center_window(900, 700)
        self.root.resizable(False, False)
        self.file_path = None
        self.questions = []
        self.current_question_index = 0
        self.incorrect_count = 0
        self.incorrect_questions = []
        self.all_questions = {}
        self.options_var = StringVar()
        self.options = []

        self.init_ui()

    def center_window(self, width, height):
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width // 2) - (width // 2)
        y = (screen_height // 2) - (height // 2)
        self.root.geometry(f"{width}x{height}+{x}+{y}")

    def init_ui(self):
        # Frame for Load File, Select Sheet, and Start Quiz buttons
        top_frame = Frame(self.root)
        top_frame.pack(pady=5)

        self.file_label = Label(
            top_frame, text="No file selected", font=("Cambria", 12, "bold")
        )
        self.file_label.pack(side=LEFT, padx=5)

        self.load_button = Button(
            top_frame, text="Load File", command=self.choose_file, font=("Cambria", 12)
        )
        self.load_button.pack(side=LEFT, padx=5)

        self.sheet_label = Label(
            top_frame, text="Select a sheet:", font=("Cambria", 12, "italic")
        )
        self.sheet_label.pack(side=LEFT, padx=5)

        self.sheet_var = StringVar(self.root)
        self.sheet_menu = OptionMenu(top_frame, self.sheet_var, [])
        self.sheet_menu.pack(side=LEFT, padx=5)

        self.num_questions_label = Label(
            top_frame, text="Num. ques:", font=("Cambria", 12)
        )
        self.num_questions_label.pack(side=LEFT, padx=5)

        self.num_questions_entry = Entry(
            top_frame, font=("Cambria", 12), width=5
        )
        self.num_questions_entry.pack(side=LEFT, padx=5)

        self.total_questions_label = Label(
            top_frame, text="Total: 0", font=("Cambria", 12)
        )
        self.total_questions_label.pack(side=LEFT, padx=5)

        # Frame for question range and shuffle option
        middle_frame = Frame(self.root)
        middle_frame.pack(pady=5)

        self.start_question_label = Label(
            middle_frame, text="From:", font=("Cambria", 12)
        )
        self.start_question_label.pack(side=LEFT, padx=5)

        self.start_question_entry = Entry(
            middle_frame, font=("Cambria", 12), width=5
        )
        self.start_question_entry.pack(side=LEFT, padx=5)

        self.end_question_label = Label(
            middle_frame, text="To:", font=("Cambria", 12)
        )
        self.end_question_label.pack(side=LEFT, padx=5)

        self.end_question_entry = Entry(
            middle_frame, font=("Cambria", 12), width=5
        )
        self.end_question_entry.pack(side=LEFT, padx=5)

        self.shuffle_var = BooleanVar(value=True)  # Default is True
        self.shuffle_check = Checkbutton(
            middle_frame, text="Shuffle Ques", variable=self.shuffle_var, font=("Cambria", 12)
        )
        self.shuffle_check.pack(side=LEFT, padx=5)

        self.start_button = Button(
            middle_frame, text="Start Quiz", command=self.start_quiz, font=("Cambria", 12)
        )
        self.start_button.pack(side=LEFT, padx=5)

        self.question_label = Label(
            self.root,
            text="",
            wraplength=800,
            font=("Cambria", 14, "bold"),
            justify="left",
        )
        self.question_label.pack(pady=10)

        self.options_frame = Frame(self.root)
        self.options_frame.pack(pady=10)

        # Frame for Submit and Next buttons
        bottom_frame = Frame(self.root)
        bottom_frame.pack(pady=5)

        self.submit_button = Button(
            bottom_frame,
            text="Submit",
            command=self.submit_answer,
            font=("Cambria", 12),
        )
        self.submit_button.pack(side=LEFT, padx=5)

        self.next_button = Button(
            bottom_frame,
            text="Next",
            command=self.next_question,
            state=DISABLED,
            font=("Cambria", 12),
        )
        self.next_button.pack(side=LEFT, padx=5)

        self.retry_button = Button(
            bottom_frame,
            text="Retry Incorrect Questions",
            command=self.retry_incorrect,
            state=DISABLED,
            font=("Cambria", 12),
        )
        self.retry_button.pack(side=LEFT, padx=5)

        self.result_label = Label(self.root, text="")
        self.result_label.pack(pady=10)


    def choose_file(self):
        self.file_path = filedialog.askopenfilename(
            filetypes=[("Word and Excel files", "*.docx *.xlsx")]
        )
        if self.file_path:
            self.file_label.config(text=os.path.basename(self.file_path))
            self.load_questions()

    def load_questions(self):
        self.all_questions = {}
        if self.file_path.endswith(".xlsx"):
            self.load_sheets()
        elif self.file_path.endswith(".docx"):
            self.load_word_document()
            total_questions = sum(len(all_questions) for all_questions in self.all_questions.values())
            self.total_questions_label.config(text=f"Total: {total_questions}")

    def load_sheets(self):
        xls = pd.ExcelFile(self.file_path)
        self.all_questions = {
            sheet: self.process_sheet(sheet, self.file_path)
            for sheet in xls.sheet_names
        }
        self.sheet_var.set("")
        menu = self.sheet_menu["menu"]
        menu.delete(0, "end")
        for sheet in xls.sheet_names:
            menu.add_command(
                label=sheet, command=lambda value=sheet: self.on_sheet_select(value)
            )

    def on_sheet_select(self, value):
        self.sheet_var.set(value)
        total_questions = len(self.all_questions[value])
        self.total_questions_label.config(text=f"Total: {total_questions}")


    def process_sheet(self, sheet, file_path):
        df = pd.read_excel(file_path, sheet)
        wb = load_workbook(file_path, data_only=True)
        ws = wb[sheet]

        questions = []
        current_question = None
        choices = []
        correct_answer = None

        for index, row in df.iterrows():
            cell_index = index + 2  # Adjust for zero-based index and header row
            if not pd.isnull(row.iloc[0]):  # New question
                if current_question:  # Save the previous question
                    questions.append(
                        {
                            "question": current_question,
                            "choices": [
                                choice for choice in choices if not pd.isnull(choice)
                            ],
                            "correct_answer": correct_answer,
                        }
                    )
                    current_question = None
                    choices = []
                    correct_answer = None

                current_question = row.iloc[1]
            if not pd.isnull(row.iloc[2]):
                choices.append(row.iloc[2])
                if self.get_correct_answer(
                    ws, cell_index, 3
                ):  # Check the color of the answer cell
                    correct_answer = len(choices) - 1  # Update the correct answer index

        if current_question:  # Save the last question
            questions.append(
                {
                    "question": current_question,
                    "choices": [choice for choice in choices if not pd.isnull(choice)],
                    "correct_answer": correct_answer,
                }
            )

        return questions

    def get_correct_answer(self, ws, row, col):
        cell = ws.cell(row=row, column=col)
        return (
            cell.fill.bgColor.index != "00000000"
            or cell.fill.fgColor.index != "00000000"
        )

    def load_word_document(self):
        doc = Document(self.file_path)
        questions = []
        current_question = None
        choices = []
        correct_answer = None

        def has_shading_or_highlight(run):
            if run.font.highlight_color is not None:
                return True
            if run._element is not None and run._element.rPr is not None:
                shading = run._element.rPr.find(qn("w:shd"))
                if shading is not None:
                    return True
            return False

        bool_ques = False
        for para in doc.paragraphs:
            # Check if the paragraph is part of a question (bold text)
            if all(run.bold for run in para.runs):
                if bool_ques:
                    questions.append(
                        {
                            "question": current_question,
                            "choices": choices,
                            "correct_answer": correct_answer,
                        }
                    )
                    current_question = None
                    choices = []
                    correct_answer = None
                    bool_ques = False
                if current_question is None:
                    current_question = para.text.strip()
                else:
                    current_question += "\n" + para.text.strip()
            else:
                bool_ques = True
                # Check if the choice has shading or highlight
                if any(has_shading_or_highlight(run) for run in para.runs):
                    correct_answer = len(choices)
                choices.append(para.text.strip())

        if current_question:  # Save the last question
            questions.append(
                {
                    "question": current_question,
                    "choices": choices,
                    "correct_answer": correct_answer,
                }
            )

        self.all_questions["Word Document"] = questions
        self.sheet_var.set("Word Document")

    def start_quiz(self):
        selected_sheet = self.sheet_var.get()
        if not selected_sheet:
            messagebox.showerror("Error", "Please select a sheet")
            return
        self.questions = self.all_questions[selected_sheet]

        # Get number of questions to ask
        # num_questions = len(self.questions)
        num_questions_str = self.num_questions_entry.get()
        if num_questions_str.isdigit():
            num_questions = int(num_questions_str)
        else:
            num_questions = len(self.questions)
        
        # Get start and end questions
        start_question_str = self.start_question_entry.get()
        end_question_str = self.end_question_entry.get()
        if start_question_str.isdigit():
            start_question = int(start_question_str) - 1  # Adjust for zero-based index
        else:
            start_question = 0

        if end_question_str.isdigit():
            end_question = int(end_question_str)
        else:
            end_question = len(self.questions)
        
        if start_question < 0 or end_question > len(self.questions) or start_question >= end_question:
            messagebox.showerror("Error", "Invalid question range")
            return
        
        self.questions = self.questions[start_question:end_question]
        
        if num_questions > len(self.questions):
            num_questions = len(self.questions)

        self.questions = self.questions[:num_questions]  # Select the required number of questions
        
        if self.shuffle_var.get():  # Check if shuffle is enabled
            self.questions = random.sample(self.questions, len(self.questions))
        
        self.current_question_index = 0
        self.incorrect_count = 0
        self.incorrect_questions = []
        self.display_question()

    def display_question(self):
        for widget in self.options_frame.winfo_children():
            widget.destroy()

        if self.current_question_index >= len(self.questions):
            self.show_result()
            return

        q = self.questions[self.current_question_index]
        self.question_label.config(
            text=f"Q{self.current_question_index + 1}: {q['question']}",
            font=("Cambria", 14, "bold"),
            anchor="w",
            justify="left",
            wraplength=800,
        )
        self.options_var.set(None)  # Reset the options variable to None

        choices = list(enumerate(q["choices"]))

        self.options = []
        for i, (original_idx, choice) in enumerate(choices):
            frame = Frame(self.options_frame)
            frame.pack(fill="x", anchor="w", padx=5)

            label = Label(
                frame, text=f"{chr(65 + i)}.", font=("Cambria", 14, "bold"), bg="white"
            )
            label.pack(side="left")

            rb = Radiobutton(
                frame,
                text=choice,
                font=("Cambria", 14),
                variable=self.options_var,
                value=str(original_idx),
                anchor="w",
                wraplength=800,
                justify="left",
                indicatoron=False,
                # height=3,
                width=810,
                selectcolor="#C6FFFD",
                bg="white",
            )
            rb.pack(side="left", fill="x")

            self.options.append(rb)

        self.submit_button.config(state=NORMAL)
        self.next_button.config(state=DISABLED)
        self.retry_button.config(state=DISABLED)

    def submit_answer(self):
        selected_option = self.options_var.get()
        correct_option = self.questions[self.current_question_index]["correct_answer"]

        if (selected_option is None) or (selected_option == "") or (selected_option == "None"):
            self.result_label.config(
                text=f"Incorrect! The correct answer is {chr(65 + correct_option)}" if correct_option is not None else "Incorrect! There is no correct answer for this question.",
                fg="red",
                font=("Cambria", 14, "bold"),
            )
            self.incorrect_count += 1
            self.incorrect_questions.append(self.questions[self.current_question_index])
        else:
            selected_option = int(selected_option)
            if selected_option == correct_option:
                self.result_label.config(
                    text="Correct!", fg="green", font=("Cambria", 14, "bold")
                )
            else:
                self.result_label.config(
                    text=f"Incorrect! The correct answer is {chr(65 + correct_option)}" if correct_option is not None else "Incorrect! There is no correct answer for this question.",
                    fg="red",
                    font=("Cambria", 14, "bold"),
                )
                self.incorrect_count += 1
                self.incorrect_questions.append(self.questions[self.current_question_index])

        self.submit_button.config(state=DISABLED)
        self.next_button.config(state=NORMAL)

    def next_question(self):
        self.result_label.config(text="")
        self.current_question_index += 1
        self.display_question()

    def show_result(self):
        total_questions = len(self.questions)
        correct_answers = total_questions - self.incorrect_count
        score_percentage = (correct_answers / total_questions) * 100

        self.question_label.config(
            text=f"Quiz Completed!\n\nTotal Questions: {total_questions}\nCorrect Answers: {correct_answers}\nScore: {score_percentage:.2f}%", font=('Cambria', 14, 'bold')
        )

        for widget in self.options_frame.winfo_children():
            widget.destroy()

        self.submit_button.config(state=DISABLED)
        self.next_button.config(state=DISABLED)

        if self.incorrect_questions:
            self.retry_button.config(state=NORMAL)

    def retry_incorrect(self):
        self.questions = self.incorrect_questions
        if self.shuffle_var.get():  # Check if shuffle is enabled
            random.shuffle(self.questions)
        self.current_question_index = 0
        self.incorrect_count = 0
        self.incorrect_questions = []
        self.display_question()



if __name__ == "__main__":
    root = Tk()
    ico = Image.open("icon.jpeg")
    photo = ImageTk.PhotoImage(ico)
    root.iconphoto(False, photo)
    app = QuizApp(root)
    root.mainloop()
