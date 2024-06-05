import pandas as pd
import random
import os
from openpyxl import load_workbook
from tkinter import *
from tkinter import filedialog, messagebox


class QuizApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Quiz Application")
        self.center_window(650, 600)
        self.root.resizable(False, False)
        self.file_path = None
        self.questions = []
        self.current_question_index = 0
        self.incorrect_count = 0
        self.all_questions = {}
        self.options_var = StringVar()
        self.options = []

        self.init_ui()

    def center_window(self, width, height):
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width // 2) - (width // 2)
        y = (screen_height // 2) - (height // 2)
        self.root.geometry(f"{width}x{height}+{x}+{y}")

    def init_ui(self):
        self.file_label = Label(self.root, text="No file selected")
        self.file_label.pack(pady=10)

        self.load_button = Button(self.root, text="Load File", command=self.choose_file)
        self.load_button.pack(pady=5)

        self.sheet_label = Label(self.root, text="Select a sheet:")
        self.sheet_label.pack(pady=5)

        self.sheet_var = StringVar(self.root)
        self.sheet_menu = OptionMenu(self.root, self.sheet_var, [])
        self.sheet_menu.pack(pady=5)

        self.start_button = Button(
            self.root, text="Start Quiz", command=self.start_quiz
        )
        self.start_button.pack(pady=10)

        self.question_label = Label(
            self.root,
            text="",
            wraplength=600,
            font=("Cambria", 12, "bold"),
            justify="left",
        )
        self.question_label.pack(pady=10)

        self.options_frame = Frame(self.root)
        self.options_frame.pack(pady=10)

        self.submit_button = Button(
            self.root, text="Submit", command=self.submit_answer
        )
        self.submit_button.pack(pady=5)

        self.next_button = Button(
            self.root, text="Next", command=self.next_question, state=DISABLED
        )
        self.next_button.pack(pady=5)

        self.result_label = Label(self.root, text="")
        self.result_label.pack(pady=10)

    def choose_file(self):
        self.file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx")]
        )
        if self.file_path:
            self.file_label.config(text=os.path.basename(self.file_path))
            self.load_sheets()

    def load_sheets(self):
        xls = pd.ExcelFile(self.file_path)
        self.all_questions = {
            sheet: self.process_sheet(sheet, self.file_path)
            for sheet in xls.sheet_names
        }
        self.sheet_var.set("")
        menu = self.sheet_menu["menu"]
        menu.delete(0, "end")
        for sheet in xls.sheet_names:
            menu.add_command(
                label=sheet, command=lambda value=sheet: self.sheet_var.set(value)
            )

    def process_sheet(self, sheet, file_path):
        df = pd.read_excel(file_path, sheet)
        wb = load_workbook(file_path, data_only=True)
        ws = wb[sheet]

        questions = []
        current_question = None
        choices = []
        correct_answer = None

        for index, row in df.iterrows():
            cell_index = index + 2  # Adjust for zero-based index and header row
            if not pd.isnull(row[0]):  # New question
                if current_question:  # Save the previous question
                    questions.append(
                        {
                            "question": current_question,
                            "choices": choices,
                            "correct_answer": correct_answer,
                        }
                    )
                    current_question = None
                    choices = []
                    correct_answer = None

                current_question = row[1]
            choices.append(row[2])
            if self.get_correct_answer(
                ws, cell_index, 3
            ):  # Check the color of the answer cell
                correct_answer = len(choices) - 1  # Update the correct answer index

        if current_question:  # Save the last question
            questions.append(
                {
                    "question": current_question,
                    "choices": choices,
                    "correct_answer": correct_answer,
                }
            )

        return questions

    def get_correct_answer(self, ws, row, col):
        cell = ws.cell(row=row, column=col)
        return (
            cell.fill.bgColor.index != "00000000"
            or cell.fill.fgColor.index != "00000000"
        )

    def start_quiz(self):
        selected_sheet = self.sheet_var.get()
        if not selected_sheet:
            messagebox.showerror("Error", "Please select a sheet")
            return
        self.questions = self.all_questions[selected_sheet]
        random.shuffle(self.questions)
        self.current_question_index = 0
        self.incorrect_count = 0
        self.display_question()

    def display_question(self):
        for widget in self.options_frame.winfo_children():
            widget.destroy()

        if self.current_question_index >= len(self.questions):
            self.show_result()
            return

        q = self.questions[self.current_question_index]
        self.question_label.config(
            text=f"Q{self.current_question_index + 1}: {q['question']}"
        )
        self.options_var.set(None)

        self.options = []
        for i, choice in enumerate(q["choices"]):
            rb = Radiobutton(
                self.options_frame,
                text=f"{chr(65 + i)}. {choice}",
                variable=self.options_var,
                value=str(i),
                anchor="w",
                wraplength=600,
                justify="left",
            )
            rb.pack(fill="x", padx=20)
            self.options.append(rb)

        self.submit_button.config(state=NORMAL)
        self.next_button.config(state=DISABLED)

    def submit_answer(self):
        if self.options_var.get() == "":
            messagebox.showwarning("Warning", "Please select an answer")
            return

        selected_option = int(self.options_var.get())
        correct_option = self.questions[self.current_question_index]["correct_answer"]

        if selected_option == correct_option:
            self.result_label.config(text="Correct!", fg="green")
        else:
            self.result_label.config(
                text=f"Incorrect! The correct answer is {chr(65 + correct_option)}",
                fg="red",
            )
            self.incorrect_count += 1

        self.submit_button.config(state=DISABLED)
        self.next_button.config(state=NORMAL)

    def next_question(self):
        self.result_label.config(text="")
        self.current_question_index += 1
        self.display_question()

    def show_result(self):
        self.question_label.config(text="")
        self.result_label.config(
            text=f"You got {self.incorrect_count} out of {len(self.questions)} questions incorrect.",
            fg="blue",
        )


if __name__ == "__main__":
    root = Tk()
    app = QuizApp(root)
    root.mainloop()
