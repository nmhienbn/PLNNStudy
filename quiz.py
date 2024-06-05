import pandas as pd
import random
import os
from openpyxl import load_workbook


class color:
    PURPLE = "\033[95m"
    CYAN = "\033[96m"
    DARKCYAN = "\033[36m"
    BLUE = "\033[94m"
    GREEN = "\033[92m"
    YELLOW = "\033[93m"
    RED = "\033[91m"
    BOLD = "\033[1m"
    UNDERLINE = "\033[4m"
    END = "\033[0m"


import os
from glob import glob


def choose_file(directory="."):
    """
    Displays all XLSX, DOCX, XLS, and DOC files in a directory and prompts the user to choose one.

    Args:
        directory (str, optional): The directory to search for files. Defaults to "." (current directory).

    Returns:
        str: The path to the chosen file or None if the user cancels the selection.
    """

    # Get all files with supported extensions
    supported_extensions = ".xlsx"
    supported_files = [
        f
        for f in glob(os.path.join(directory, "*"))
        if os.path.splitext(f)[1].lower() in supported_extensions
    ]

    # Check if there are any supported files
    if not supported_files:
        print(
            f"No files with extensions {', '.join(supported_extensions)} found in {directory}."
        )
        return None

    # Print file list for user selection
    print("Available files:")
    for i, filename in enumerate(supported_files):
        print(f"{i+1}. {filename}")

    # Get user input for file selection
    while True:
        choice = input(
            "Enter the number of the file you want to choose, or 'q' to cancel: "
        )
        if choice.lower() == "q":
            print("Selection cancelled.")
            return None

        try:
            choice_index = int(choice) - 1
            if 0 <= choice_index < len(supported_files):
                return supported_files[choice_index]
            else:
                print(
                    "Invalid choice. Please enter a valid file number or 'q' to cancel."
                )
        except ValueError:
            print("Invalid input. Please enter a number or 'q'.")


def process_sheet(sheet, file_path):
    df = pd.read_excel(file_path, sheet)
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet]

    questions = []
    current_question = None
    choices = []
    correct_answer = None

    for index, row in df.iterrows():
        cell_index = index + 2  # Adjust for zero-based index and header row
        if not pd.isnull(row[0]):  # New question
            if current_question:  # Save the previous question
                questions.append(
                    {
                        "question": current_question,
                        "choices": choices,
                        "correct_answer": correct_answer,
                    }
                )
                current_question = None
                choices = []
                correct_answer = None

            current_question = row[1]
        choices.append(row[2])
        if get_correct_answer(ws, cell_index, 3):  # Check the color of the answer cell
            correct_answer = len(choices) - 1  # Update the correct answer index

    if current_question:  # Save the last question
        questions.append(
            {
                "question": current_question,
                "choices": choices,
                "correct_answer": correct_answer,
            }
        )

    return questions


def get_correct_answer(ws, row, col):
    cell = ws.cell(row=row, column=col)
    return (
        cell.fill.bgColor.index != "00000000" or cell.fill.fgColor.index != "00000000"
    )


def ask_questions(questions):
    incorrect_count = 0
    random.shuffle(questions)  # Shuffle the order of questions

    for i, q in enumerate(questions):
        print(
            color.BOLD
            + color.UNDERLINE
            + f"\nQ{i+1}:"
            + color.END
            + f" {q['question']}"
        )

        choices = q["choices"]
        correct_answer_index = q["correct_answer"]

        # Shuffle choices and track the index of the correct answer
        indexed_choices = list(enumerate(choices))
        random.shuffle(indexed_choices)

        new_correct_answer_index = None
        for idx, (original_idx, choice) in enumerate(indexed_choices):
            print(color.BOLD + f"   {chr(65+idx)}." + color.END + f" {choice}")
            if original_idx == correct_answer_index:
                new_correct_answer_index = idx

        answer = (
            input(color.BOLD + color.UNDERLINE + "Your answer:" + color.END + " ")
            .strip()
            .upper()
        )
        valid_choices = [chr(65 + k) for k in range(len(choices))]

        while answer not in valid_choices:
            answer = (
                input(
                    "Invalid choice. Please enter a valid option ("
                    + color.BOLD
                    + color.UNDERLINE
                    + f"{', '.join(valid_choices)}"
                    + color.END
                    + "): "
                )
                .strip()
                .upper()
            )

        if ord(answer) - 65 == new_correct_answer_index:
            print(color.BOLD + color.GREEN + "CORRECT!" + color.END)
        elif new_correct_answer_index == None:
            print(
                color.BOLD
                + color.RED
                + "No answer in file."
                + color.END
            )
            incorrect_count += 1
        else:
            print(
                color.BOLD
                + color.RED
                + f"INCORRECT! The correct answer is {chr(65 + new_correct_answer_index)}"
                + color.END
            )
            incorrect_count += 1

        input(color.YELLOW + "Press any key to continue..." + color.END)

    print(
        color.BOLD
        + color.BLUE
        + f"\nYou got {incorrect_count} out of {len(questions)} questions incorrect.\n\n"
        + color.END
    )


def main():
    file_path = choose_file()

    xls = pd.ExcelFile(file_path)

    all_questions = {
        sheet: process_sheet(sheet, file_path) for sheet in xls.sheet_names
    }

    sheets = list(all_questions.keys())
    while True:
        print("Available quizzes:")
        for i, sheet in enumerate(sheets):
            print(f"{i + 1}. {sheet}")

        sheet_index = input(
            "Select a quiz by entering the corresponding number: "
        ).strip()
        while not sheet_index.isdigit() or int(sheet_index) not in range(
            1, len(sheets) + 1
        ):
            sheet_index = input(
                f"Invalid choice. Please enter a number between 1 and {len(sheets)}: "
            ).strip()

        selected_sheet = sheets[int(sheet_index) - 1]
        questions = all_questions[selected_sheet]

        print(f"\n=== {selected_sheet} ===")

        ask_questions(questions)


if __name__ == "__main__":
    main()
